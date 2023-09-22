import json

import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials

from base_object_presenter.services import BaseServicesPresenter
from product.models import Product
from project import settings
from .models import StaffModelPresenter

from openpyxl import Workbook
from openpyxl.drawing.image import Image


class StaffServicesPresenter(BaseServicesPresenter):
    def __init__(self):
        self.model_presenter = StaffModelPresenter()
        super().__init__()

        self.CREDENTIALS_FILE = 'site_settings/creds.json'
        self.spreadsheet_id = '1phdFtzmKHjXv0yxwD39b9kVFVRLstsdiReuK8Fa_tPk'

        self.credentials = ServiceAccountCredentials.from_json_keyfile_name(
            self.CREDENTIALS_FILE,
            ['https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive'])
        self.httpAuth = self.credentials.authorize(httplib2.Http())
        self.service = apiclient.discovery.build('sheets', 'v4', http=self.httpAuth)

    def add_order(self, data):
        data["product"] = json.loads(data["product"])

        company_name = data["company_name"]
        product_count = data["count"]
        product_name = data["product"]["name"]
        product_code = data["product"]["code"]
        product_image = f'= IMAGE("http://{settings.SITE_DOMAIN + data["product"]["image"]}"; 2)'
        product_price = data["product"]["price"]
        order_date = data["date"]

        if data["product"]["category_name"] == "Товары со склада":
            sheetname = "Заказы (Китай)"
        else:
            sheetname = "Заказы (Базар)"

        response = self.service.spreadsheets().values().get(
            spreadsheetId=self.spreadsheet_id,
            range=sheetname
        ).execute()

        values = response.get('values', [])

        if values:
            company_names = []

            for row in values:
                try:
                    company_names.append(row[0])
                except:
                    pass

            if company_name in company_names:
                index = company_names.index(company_name)
                updated = False

                if product_code:
                    for i in range(index, len(company_names)):
                        if company_names[i] == company_name:
                            if values[i][3] == product_code and values[i][6] == order_date:
                                values[i][1] = int(values[i][1]) + product_count
                                updated = True
                                index = i
                                break

                if not updated:
                    self.service.spreadsheets().values().append(
                        spreadsheetId=self.spreadsheet_id,
                        range=sheetname,
                        body={'values': [
                            [company_name,
                            product_count,
                            product_name,
                            product_code,
                            product_image,
                            product_price,
                            order_date]
                        ]},
                        valueInputOption='USER_ENTERED',
                    ).execute()
                    return

                update_range = f'{sheetname}!B{index + 1}'
                request_body = {'values': [
                    [values[index][1]]
                ]}

                self.service.spreadsheets().values().update(
                    spreadsheetId=self.spreadsheet_id,
                    range=update_range,
                    body=request_body,
                    valueInputOption='USER_ENTERED'
                ).execute()
            else:
                append_values = [[
                    company_name,
                    product_count,
                    product_name,
                    product_code,
                    product_image,
                    product_price,
                    order_date
                ]]
                self.service.spreadsheets().values().append(
                    spreadsheetId=self.spreadsheet_id,
                    range=sheetname,
                    body={'values': append_values},
                    valueInputOption='USER_ENTERED',
                ).execute()

    def get_products_in_excel(self, start, end):
        wb = Workbook()

        sheet = wb.active
        sheet.append(
            ["ID", "Категория", "Названия", "Описания", "Цена", "В наличии", "Артикул", "Номер поставщика", "Высота", "Ширина", "Длина", "Количество", "Фото"]
        )

        products = Product.objects.prefetch_related("images", "category").order_by("-id").all()[int(start):int(end)]

        for product in products:
            sheet.append([product.id, product.category.name, product.name, product.description, product.price, product.is_available, product.code, product.vendor_number, product.height, product.width, product.length, product.count])

            try:
                image_url = product.images.first().image.url
                #image_path = product.images.filter(default=True).first().image.path
                sheet[f"M{sheet.max_row}"].value = 'https://kassym.com' + image_url
                sheet[f"M{sheet.max_row}"].hyperlink = 'https://kassym.com' + image_url
                continue

                if '.webp' in image_path:
                    sheet[f"M{sheet.max_row}"] = "https://kassym.com" + image_url
                else:
                    img = Image(image_path)
                    img.width = 100
                    img.height = 100

                    sheet.add_image(img, f"M{sheet.max_row}")
            except Exception as e:
                pass

        wb.save("output.xlsx")
        return "output.xlsx"
